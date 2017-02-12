#-*- coding: utf-8 -*-
# -------------------------------------------------------------------------------
# Name:        ParseExcel.py
# Purpose:
#
# Author:      Oreo
#
# Created:     29-01-2017
# Copyright:   (c) Oreo 2017
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import os
import openpyxl
import xlrd

# from openpyxl import load_workbook     <-- openpyxl를 사용하기 위함.
# wb = load_workbook(엑셀파일이름)            <-- 엑셀 파일이름을 넣는다.
# sheetList = wb.get_sheet_names()        <-- sheet 이름들을 리스트로  얻는다.
# sheet = wb.get_sheet_by_name(sheetList[i])   <-- 원하는 sheet 의 포인트를 얻는다.
# cellValue = sheet.cell(셀위치).value

# http://dart.fss.or.kr/pdf/download/excel.do?rcp_no=20161114002564&lang=ko
def parseFSsheet(rcp_no):
    # ex) 20161114002564_ko.xls
    fsReportFileName = rcp_no +"_ko.xls"
    fsReportFullPath = os.path.join("FS_xls", fsReportFileName)
    workbook = xlrd.open_workbook(fsReportFullPath)
    for i in workbook.sheet_names():  # 시트목록을 출력(list형식)
        print i
    worksheet_name = workbook.sheet_by_name(u'연결 재무상태표')  # 시트이름으로 시트 가져오기
    worksheet_index = workbook.sheet_by_index(1)  # 시트번호(인덱스)로 시트 가져오기

    num_rows = worksheet_name.nrows  # 줄 수 가져오기
    num_cols = worksheet_name.ncols  # 칸 수 가져오기

    # worksheet_name.col_values()
    row_val = worksheet_name.row_values(8)  # 줄 값 가져오기(list 형식)
    cell_val = worksheet_name.cell_value(8, 2)  # 셀 값 가져오기
    print row_val

    # xlrd 시트 인덱스 시작은 0부터
    # Start from row 8,1
    # 딕셔너리에 행 index 1 값(항목이름)을 받아 비교 후 해당 딕셔너리 항목에 넣기
    for row_idx in range(8, num_rows):
        print ('-' * 40)
        print ('Row: %s' % row_idx)  # Print row number
        for col_idx in range(0, num_cols):  # Iterate
            cell_obj = worksheet_name.cell(row_idx, col_idx)  # Get cell object by row, col
            print ('Column: [%s] cell_obj: [%s]' % (col_idx, cell_obj))

    pass


def main():
    """
    # openpyxl
    wb = openpyxl.load_workbook(r'C:\Users\Oreo\Desktop\20161114002564_ko.xls')
    sheetList = wb.get_sheet_by_name()
    for i in sheetList:
        print i
    #sheet = wb.get_sheet_by_name(sheetList[i])
    #cellValue = sheet.cell().value
    """

    workbook = xlrd.open_workbook(r'C:\Users\Oreo\Desktop\20161114002564_ko.xls')
    workbook = xlrd.open_workbook(os.path.join("FS_xls", '20161114002564_ko.xls'))
    print "시트목록을 출력"
    for i in workbook.sheet_names(): # 시트목록을 출력(list형식)
        print i
    worksheet_name = workbook.sheet_by_name(u'연결 재무상태표')  # 시트이름으로 시트 가져오기
    worksheet_index = workbook.sheet_by_index(1)  # 시트번호(인덱스)로 시트 가져오기

    num_rows = worksheet_name.nrows  # 줄 수 가져오기
    num_cols = worksheet_name.ncols  # 칸 수 가져오기

    # 시트 인덱스 시작은 0부터
    # row 8,1
    # worksheet_name.col_values()
    row_val = worksheet_name.row_values(8)  # 줄 값 가져오기(list형식)
    cell_val = worksheet_name.cell_value(9, 2)  # 셀 값 가져오기
    # print "cell row 9 value 출력"
    # print row_val

    # test run parseFSsheet
    # example rcp_no : 20161114002564
    parseFSsheet('20161114002564')

    # Ref code
    # 엑셀에 있는 값 가져오기
    #workbook = xlrd.open_workbook('example.xls')
    #worksheet = workbook.sheet_by_index(0)
    #nrows = sheet.nrows

    #row_val = []
    #for row_num in range(nrows):
    #    row_val.append(worksheet.row_values(row_num))

    pass

if __name__ == '__main__':
    main()
